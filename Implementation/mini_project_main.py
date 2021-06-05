# This file contains main source code
"""
    * [File: mini_project_main.py]
    * [Author: Debashish Dash]
    * [Date: 03-06-2021]
    *
    * [@copyright (c) 2021]
    *
"""

import os
import sys
from openpyxl import load_workbook #, workbook


class SystemOperations:
    """[This class contains the method to check the validity of an userinput
        and all the system operations like clear screen and exit screen]
    """
    @staticmethod
    def screen_clear():                                 # static method
        """
        clears the screen, works in both the os
        """
    # for mac and linux(here, os.name is 'posix')
    if os.name == 'posix':
        _ = os.system('clear')
    else:
        # for windows platfrom
        _ = os.system('cls')
    # End of Function


    @staticmethod                                       # static method
    def system_exit():
        """
            [Exits from the code]
        """
        sys.exit("\n\nThank You for visiting... \n\n")
    # End of Function

    def validate_ip(self, user_ip, flatten_ps):
        """[This method checks the validity the choices entered by the user]

        Args:
            user_ip ([type: char]): [user choice]
            flatten_ps ([type: list]): [available options]

        Returns:
            [type: boolean]: [True if present and false if not present in the list]
        """
        int_user_ip=int(user_ip)
        if int_user_ip in flatten_ps:
            return True
        return False
    # End of function


class MyExcel(SystemOperations):                    # inharitance
    """[This class contains all the basic methos for getting values from the excel sheet]

    Inharitance:
        systemOperations ([type: class]): [MyExcel class inherits the systemOperations class]
    """
    def __init__(self,work_book):
        self.work_book = work_book
    # End of const

    def get_ps_number(self,ip_ws):
        """[fetch the data from the first column of the excel]

        Args:
            ip_ws ([type: object]): [active worksheet]

        Returns:
            [type: list]: [data of the 1st column in the excel sheet, return as list]
        """
        ps_nums=[]
        for row in ip_ws.iter_rows(min_row=1, max_col=1, max_row=16, values_only=True):
            ps_nums.append(list(row))
        return ps_nums
    # End Of Function

    def flatten_ps_number(self,ps_list):
        """[so this function flatten the list in to 1-D]

        Args:
            ps_list ([type: list]): [It is a 2-D list]

        Returns:
            [type: list]: [Converted 1-D form of the 2-D list is returned]
        """
        flat_ps = []
        # iterating over the list that contains ps nums
        for ps_num in ps_list:
            #appending ps number to the flat list
            flat_ps += ps_num
        return flat_ps
    # End of function


    def display_ps_number(self,flat_ps):
        """[prints the ps ids in the terminal]

        Args:
            flat_ps ([type: list]): [1-d converted form of the 2-D list data]
        """
        for ps_num in flat_ps:
            print(ps_num)
    # End of Function


    def display_sheets_names(self):
        """[stores and Displays the names of the sheets]
        """
        #self.work_book=work_book
        work_sheets = self.work_book.sheetnames       #storing the worksheet names of the workbook
        opt=1
        for sheet in work_sheets:
            print("\npress ",opt," for ",sheet)
            opt+=1
    # End of Function
# End Of class myExcel

class MyExcelOperations:
    """[This class contains all the methods used to fetch the data]
    """
    def __init__(self,work_book,work_sheet, ls_sheet_name):
        self.work_book = work_book
        self.work_sheet = work_sheet
        self.my_excel_obj_in = MyExcel(work_book)      # object of this class
        self.ls_sheet_name = ls_sheet_name
    # End of Const


    def get_rows(self,work_sheet,ps_num):
        """[returns the row index object]

        Args:
            work_sheet ([type: worksheet object]): [choosen work sheet]
            ps_num ([type: int]): [number entered by user]

        Returns:
            [type: object]: [returns the index in object form]
        """
        self.work_sheet=work_sheet
        for row in work_sheet.rows:
            if row[0].value == ps_num:
                return row
    # End of function


    #var = get_rows(wb['Sem'], 99004351)
    def get_data(self,row_index):
        """[gets data from the index]

        Args:
            row_index ([type: object]): [takes object of row number as input]
        """
        index_data=[]
        for index in range(1,20):
            index_data.append(row_index[index].value)
        print(index_data)
        print("\n")
        print("*"*125)
        print("\t\t\tThe Fetched data is also present in '|>>>>> Outputs.xlsx <<<<<|' file")
        print("*"*125)
    # End of function

    def write_excel(self, data, output_file, entered_ps_number, ls_sheet_name):
        """[This function writes the data to excel sheet]

        Args:
            data ([type: list]): [data to be entered]
            output_file ([type: workbook object]): [object of output workbook excel file]
            entered_ps_number ([type: int]): [ps number entered by the user]
            ls_sheet_name ([type: list]): [name of the sheet]
        """
        work_s = output_file.active       # selecting the active sheet of the output work book
        ls_data=[entered_ps_number, ls_sheet_name]
        for index in range(1,20):
            ls_data.append(data[index].value)
        work_s.append(ls_data)
        output_file.save("Outputs.xlsx")
    # End of function

    def excel_operation(self,work_sheet_name):
        """[performs all the operation to fetch the data]

        Args:
            work_sheet_name ([type: object]): [selected worksheet from the user]
        """
        user_choice=" "
        while user_choice != '1':
            print("\nEnter the PS number from the below list:\n")
            print(20*'=')
            my_excel_obj_in = self.my_excel_obj_in  # object of myExcel class
            # calling functions of myExcel class
            ps_num = my_excel_obj_in.get_ps_number(work_sheet_name)
            flatten_ps = my_excel_obj_in.flatten_ps_number(ps_num)
            my_excel_obj_in.display_ps_number(flatten_ps)
            print(20*'=')
            print("\nEnter the ps number to get its data")
            print("\nOr press 0 to exit")
            print("\nOr press 1 to go to Sheet selection menu:")
            user_choice = input("\nEnter Your Choice: ")
            my_excel_obj_in.screen_clear()                           # inherited function used
            if user_choice == '0':
                my_excel_obj_in.system_exit()                        # inherited function used
            if user_choice == '1':
                break
            if user_choice == '':
                print(20*'=')
                print("Oops You have entered nothing !!")
                print("\nPlease Try again...")
                print(20*'=')
                continue
            if user_choice in str(ps_num):
                if my_excel_obj_in.validate_ip(user_choice,flatten_ps):
                    print("="*40)
                    print("\nEntered Ps is: ", user_choice, "and respective Data is::" )
                    row_index = self.get_rows(work_sheet_name, int(user_choice))
                    self.get_data(row_index)
                    output_file = load_workbook("Outputs.xlsx")     #loading the oyutput work book
                    # writing into output workbook
                    ls_sheet_name = self.ls_sheet_name
                    self.write_excel(row_index, output_file, user_choice, ls_sheet_name)
                    print("\n")
                    print("="*40)
                    print("Enter Another PS ????")
                else:
                    print(20*'=')
                    print("Oops You have entered an invalid choice !!")
                    print("\nPlease Try again...")
                    print(20*'=')
            else:
                print(20*'=')
                print("Oops You have entered an invalid choice !!")
                print("\nPlease Try again...")
                print(20*'=')
    # End of function
# End of class MyExcelOperation


class Runner(MyExcel,MyExcelOperations):        # multiple inharitance
    """[main class uses multiple inheritance]

    Args:
        MyExcel : [MyExcel class inherited]
        MyExcelOperations : [MyExcelOperation class inherited]
    """
    def __init__(self, work_book):
        self.work_book = work_book   # loading the workbook
        self.my_excel_obj = MyExcel(work_book)    #object of the MyExcel


    def run(self):
        """[this is the function that performs all the tasks in a sequence]
        """
        #ws = wb.active           # gives the active work sheets of the workbook
        #sheet_ls = work_book.sheetnames
        # print(sheet_ls)
        sheet_score = work_book['Sem']
        sheet_hobbies = work_book['Hobbies']
        sheet_cities = work_book['Cities']
        sheet_pl = work_book['PL']
        sheet_domain = work_book['Domain']
        while True:
            print("\nSelect the Sheet:")
            print(30*'=')
            self.my_excel_obj.display_sheets_names()
            print("\nOr press 0 to exit: \n")
            print(30*'=')
            sheet_choice = input("\nPlease Enter your choice:")
            self.my_excel_obj.screen_clear()           # inherited function used
            if sheet_choice == '1':
                # myExcelOperations(work_book, sheet_score)
                ls_sheet_name = work_book.sheetnames[0]
                op_sheet_score = MyExcelOperations(work_book, sheet_score, ls_sheet_name)
                op_sheet_score.excel_operation(sheet_score)
            elif sheet_choice == '2':
                # declaring object with sheet_hobbies
                ls_sheet_name = work_book.sheetnames[1]
                op_sheet_hobbies = MyExcelOperations(work_book, sheet_hobbies, ls_sheet_name)
                op_sheet_hobbies.excel_operation(sheet_hobbies)
            elif sheet_choice == '3':
                # declaring object with sheet_cities
                ls_sheet_name = work_book.sheetnames[2]
                op_sheet_cities = MyExcelOperations(work_book, sheet_cities, ls_sheet_name)
                op_sheet_cities.excel_operation(sheet_cities)
            elif sheet_choice == '4':
                # declaring object with sheet_sheet_pl
                ls_sheet_name = work_book.sheetnames[3]
                op_sheet_pl = MyExcelOperations(work_book, sheet_pl, ls_sheet_name)
                op_sheet_pl.excel_operation(sheet_pl)
            elif sheet_choice == '5':
                # declaring object with sheet_domain
                ls_sheet_name = work_book.sheetnames[4]
                op_sheet_domain = MyExcelOperations(work_book, sheet_domain, ls_sheet_name)
                op_sheet_domain.excel_operation(sheet_domain)
            elif sheet_choice == '0':
                self.my_excel_obj.system_exit()                # inherited function used
            else:
                print(30*'=')
                print("ops..!! You have entered an invalid choice")
                print("Please Try again... :(")
                print(30*'=')
                #main()
                continue
    # End of run function


if __name__ == "__main__":
    work_book = load_workbook("Advanced_python_mini_project.xlsx")   #loading work book
    main_class_obj = Runner(work_book)    # object of main class
    main_class_obj.run()
